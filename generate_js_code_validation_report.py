import json
import boto3
import io
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

logger = logging.getLogger()
logger.setLevel(logging.INFO)

s3 = boto3.client("s3")

def lambda_handler(event, context):
    logger.info(f"Received event keys: {list(event.keys())}")
    
    # Extract validation_report correctly
    validation_data = event.get("validation_report", {})
    logger.info(f"validation_data type: {type(validation_data)}")
    
    # Parse the body if it's a string
    if isinstance(validation_data, dict) and "body" in validation_data:
        body = validation_data["body"]
        logger.info(f"Found body, type: {type(body)}")
        if isinstance(body, str):
            body = json.loads(body)
        validation_report = body.get("validation_report", [])
    else:
        validation_report = validation_data.get("validation_report", [])
    
    logger.info(f"validation_report length: {len(validation_report)}")
    
    # Count total fields
    total_fields = sum(len(seg.get("fields", [])) for seg in validation_report)
    logger.info(f"Total fields to export: {total_fields}")
    
    bucket = "hl7v2autoct"
    prefix = "output/final_js_validation/"
    filename = "JS_Code_Validation_Report.xlsx"

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Validation Report"

    # Define column headers
    headers = [
        "Segment", "Data Element", "Canonical Field", "Source Field", 
        "Transformation Rules", "Source Field JS Code", "Sample Input", 
        "Expected Output", "Actual Output", "Validation Status", 
        "Validation Comments"
    ]
    ws.append(headers)
    
    # Format header row
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Populate rows - UPDATED: Single row per field
    row_count = 0
    for segment in validation_report:
        segment_name = segment.get("segment", "")
        fields = segment.get("fields", [])
        logger.info(f"Processing segment {segment_name} with {len(fields)} fields")
        
        for field in fields:
            logger.info(f"Transformation Rules for {field.get('Canonical Field')}: {field.get('Transformation Rules')}")
            # UPDATED: Each field creates exactly ONE row (no loops over sample values)
            raw_row = [
                segment_name,
                field.get("Data Element", ""),
                field.get("Canonical Field", ""),
                field.get("Source Field", ""),
                field.get("Transformation Rules", ""),
                field.get("Source Field JS Code", ""),
                field.get("Sample Input", ""),
                field.get("Expected Output", ""),
                field.get("Actual Output", ""),
                field.get("Validation Status", ""),
                field.get("Validation Comments", "")
            ]

            # Convert any list values to comma-separated strings
            safe_row = [", ".join(v) if isinstance(v, list) else v for v in raw_row]
            ws.append(safe_row)

            row_count += 1
            
            # Apply conditional formatting to Validation Status
            status_cell = ws.cell(row=ws.max_row, column=10)  # Column J (Validation Status)
            status = field.get("Validation Status", "")
            
            if status == "Pass":
                status_cell.fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
                status_cell.font = Font(color="006100")
            elif status == "Fail":
                status_cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
                status_cell.font = Font(color="9C0006")

    logger.info(f"Total rows written: {row_count}")

    # Set column widths
    column_widths = {
        'A': 12,  # Segment
        'B': 30,  # Data Element
        'C': 15,  # Canonical Field
        'D': 15,  # Source Field
        'E': 50,  # Transformation Rules
        'F': 60,  # Source Field JS Code
        'G': 40,  # Sample Input
        'H': 40,  # Expected Output
        'I': 40,  # Actual Output
        'J': 18,  # Validation Status
        'K': 50   # Validation Comments
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Enable text wrapping for all data cells
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="top")

    # Save to memory
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    # Upload to S3
    s3.put_object(
        Bucket=bucket,
        Key=prefix + filename,
        Body=file_stream.getvalue(),
        ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    
    # Generate pre-signed URL (valid for 1 hour)
    presigned_url = s3.generate_presigned_url(
        'get_object',
        Params={'Bucket': bucket, 'Key': prefix + filename},
        ExpiresIn=3600  # 1 hour
    )
    
    s3_location = f"s3://{bucket}/{prefix}{filename}"
    logger.info(f"Excel uploaded to: {s3_location}")

    return {
        "statusCode": 200,
        "s3_path": s3_location,
        "download_url": presigned_url,
        "total_segments": len(validation_report),
        "total_rows": row_count,
        "body": json.dumps({
            "message": "Excel report uploaded successfully", 
            "location": s3_location,
            "download_url": presigned_url,
            "total_segments": len(validation_report),
            "total_rows": row_count
        })
    }
