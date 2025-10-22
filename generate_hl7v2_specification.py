import json
import boto3
import io
from datetime import datetime
from collections import defaultdict
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Initialize S3 client
s3 = boto3.client("s3")

# S3 target
bucket_name = "hl7v2autoct"
output_prefix = "output/hl7v2_specifications"

def upload_to_s3(specs):
    # timestamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
    # filename = f"specs_{timestamp}.xlsx"
    filename = "HL7v2_ORM_Specification.xlsx"
    s3_key = f"{output_prefix}/{filename}"

    # Create Excel workbook
    wb = Workbook()
    ws = wb.active
    ws.title = "Specifications"

    # Define headers
    headers = [
        "Data Element", "Canonical Field", "Source Field", "Usage",
        "Fill Rate", "Min Length", "Max Length", "Transformation Rules",
        "Sample Input Values", "Expected Output"
    ]
    ws.append(headers)

    # Format header row
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    # Populate rows
    for spec in specs:
        # Convert Transformation Rules to string if it's a dict
        trans_rules = spec.get("Transformation Rules", "")
        if isinstance(trans_rules, dict):
            # Format dict as readable string
            parts = []
            for key, value in trans_rules.items():
                parts.append(f"{key}: {value}")
            trans_rules = "; ".join(parts)
        elif isinstance(trans_rules, list):
            trans_rules = "; ".join(str(r) for r in trans_rules)
        else:
            trans_rules = str(trans_rules) if trans_rules else ""

        print("trans_rules:", trans_rules)
        
        row = [
            spec.get("Data Element", ""),
            spec.get("Canonical Field", ""),
            spec.get("Source Field", ""),
            spec.get("Usage", ""),
            spec.get("Fill Rate", ""),
            spec.get("Min Length", ""),
            spec.get("Max Length", ""),
            trans_rules,  # Use converted string
            spec.get("Sample Input Values", ""),
            spec.get("Expected Output", "")
        ]
        ws.append(row)

    # Set column widths
    column_widths = {
        'A': 30,  # Data Element
        'B': 18,  # Canonical Field
        'C': 18,  # Source Field
        'D': 12,  # Usage
        'E': 12,  # Fill Rate
        'F': 12,  # Min Length
        'G': 12,  # Max Length
        'H': 60,  # Transformation Rules (increased width)
        'I': 50,  # Sample Input Values
        'J': 50   # Expected Output
    }
    
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # Freeze header row
    ws.freeze_panes = 'A2'

    # Save to memory
    file_stream = io.BytesIO()
    wb.save(file_stream)
    file_stream.seek(0)

    # Upload to S3
    s3.put_object(
        Bucket=bucket_name,
        Key=s3_key,
        Body=file_stream.getvalue(),
        ContentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )

    return f"s3://{bucket_name}/{s3_key}", s3_key

def flatten_value(value):
    """
    Flatten a single value according to HL7v2 encoding rules:
    - Nested lists (components & subcomponents): use ^ for components, & for subcomponents
    - Simple lists: join with ^
    - Primitives: convert to string
    """
    print(f"Flattening value: {value}")
    print(f"Type of value: {type(value)}")
    
    # If it's a string that looks like JSON, try to parse it first
    if isinstance(value, str):
        # Check if it looks like a JSON array
        if value.strip().startswith('[') and value.strip().endswith(']'):
            try:
                value = json.loads(value)
                print(f"Parsed string to: {value}")
            except Exception as e:
                print(f"Failed to parse JSON string: {e}")
                return value  # Return as-is if parsing fails
        else:
            print(f"Value is not a list: {value}")
            return value  # Return plain string as-is
    
    if isinstance(value, list):
        print(f"Value is a list with {len(value)} items")
        parts = []
        for item in value:
            if isinstance(item, list):
                # Recursively handle nested lists - this handles subcomponents
                subparts = []
                for subitem in item:
                    if isinstance(subitem, list):
                        # Triple nested - join with & first
                        subsubcomp = "&".join(str(s) if s != "" else "" for s in subitem)
                        print(f"    Triple-nested subcomponent: {subsubcomp}")
                        subparts.append(subsubcomp)
                    else:
                        # Regular subcomponent
                        subparts.append(str(subitem) if subitem != "" else "")
                subcomp = "&".join(subparts)
                print(f"  Subcomponent: {subcomp}")
                parts.append(subcomp)
            else:
                # Component - convert to string
                comp = str(item) if item != "" else ""
                print(f"  Component: {comp}")
                parts.append(comp)
        result = "^".join(parts)
        print(f"Flattened result: {result}")
        return result
    
    print(f"Returning as string: {value}")
    return str(value) if value != "" else ""

def format_value_list(value_list):
    """
    Format a list of values (repetitions) for display.
    Each value is flattened, then joined with comma-space for readability.
    """
    print(f"Formatting value list: {value_list}")
    
    if not isinstance(value_list, (list, tuple)):
        value_list = [value_list] if value_list is not None else []
    
    formatted_values = []
    for v in value_list:
        flattened = flatten_value(v)
        if flattened:  # Only add non-empty values
            formatted_values.append(flattened)
    
    result = ", ".join(formatted_values)
    print(f"Final formatted result: {result}")
    return result

def lambda_handler(event, context):
    evaluations = event.get("evaluations", [])
    print("Received evaluations:", json.dumps(evaluations))

    all_specs = []
    grouped_specs = defaultdict(list)

    for segment_obj in evaluations:
        segment = segment_obj.get("segment")
        fields = segment_obj.get("fields", [])
        field_evals = segment_obj.get("evaluation_result", {})

        for field in fields:
            field_id = field.get("field_id")
            label = field.get("label")
            usage = field.get("usage")
            stats = field.get("stats", {})

            sample_values_raw = stats.get("all_values")
            sample_values = []
            if sample_values_raw:
                try:
                    parsed_values = json.loads(sample_values_raw)
                    if isinstance(parsed_values, (list, tuple)):
                        sample_values = parsed_values
                    else:
                        sample_values = [parsed_values]
                except Exception:
                    sample_values = [sample_values_raw]

            fill_rate = stats.get("fill_rate")
            has_stats = bool(stats) and fill_rate not in [None, "0", 0, "0.0", 0.0]

            field_eval = field_evals.get("evaluations", {}).get(field_id, {})
            rule_triggered = field_eval.get("Rule Triggered", "No")
            transformation = field_eval.get("Transformation Rules", "")
            expected_output_raw = field_eval.get("Expected Output", "")

            source_field = (
                "NA" if not has_stats else
                "Logic" if rule_triggered == "Yes" else
                field_id
            )

            expected_output_list = expected_output_raw if isinstance(expected_output_raw, list) else sample_values

            if not isinstance(expected_output_list, (list, tuple)):
                expected_output_list = [expected_output_list] if expected_output_list is not None else []

            # Format the outputs - flatten_value now handles string parsing
            print("--- Formatting Expected Output ---")
            formatted_expected_output = format_value_list(expected_output_list)
            print(f"Formatted Expected Output: {formatted_expected_output}")
            
            print("--- Formatting Sample Values ---")
            formatted_sample_values = format_value_list(sample_values)
            print(f"Formatted Sample Values: {formatted_sample_values}")

            spec_row = {
                "Data Element": label,
                "Canonical Field": field_id,
                "Source Field": source_field,
                "Usage": usage,
                "Fill Rate": fill_rate if has_stats else "",
                "Min Length": stats.get("min_length", ""),
                "Max Length": stats.get("max_length", ""),
                "Transformation Rules": transformation if transformation else "",
                "Sample Input Values": formatted_sample_values,
                "Expected Output": formatted_expected_output
            }

            all_specs.append(spec_row)
            grouped_specs[segment].append(spec_row)

    s3_path, s3_key = upload_to_s3(all_specs)

    # Generate a signed download URL
    download_url = s3.generate_presigned_url(
        ClientMethod='get_object',
        Params={
            'Bucket': bucket_name,
            'Key': s3_key
        },
        ExpiresIn=3600  # 1 hour
    )


    return {
        "statusCode": 200,
        "specs": [
            {
                "segment": segment,
                "fields": fields
            }
            for segment, fields in grouped_specs.items()
        ],
        "s3_path": s3_path,
        "download_url": download_url
    }